"""
Kalshi Candlestick Backtester
==============================
Eliminates look-ahead bias by walking through hourly price candles
chronologically and simulating signals exactly as they would fire
in real trading — BEFORE the market resolves.

How it works:
  1. Pull settled markets (so we know the actual outcome)
  2. For each market, fetch its full hourly candlestick history
  3. Walk candles in time order — stop BEFORE the final hours
  4. Apply each strategy signal at each candle using only
     data available at that point in time
  5. Record the FIRST signal trigger per market per strategy
  6. P&L = entry price at signal time vs actual settlement result

This is the correct way to backtest — no peeking at the answer.

Usage:
    python kalshi_backtest_v2.py                  # 100 markets (fast)
    python kalshi_backtest_v2.py --markets 300    # more markets, slower
    python kalshi_backtest_v2.py --save           # saves Excel report
    python kalshi_backtest_v2.py --verbose        # show candle-level detail

Requirements:
    pip install requests tabulate colorama openpyxl
"""

import requests
import time
import argparse
from datetime import datetime, timezone
from datetime import datetime, timezone
from dataclasses import dataclass, field
from collections import defaultdict
from typing import Optional
from tabulate import tabulate
from colorama import init, Fore
init(autoreset=True)

BASE_URL = "https://api.elections.kalshi.com/trade-api/v2"

# ── Constants ────────────────────────────────────────────────────────────────
FEE_PER_CONTRACT    = 0.00   # Kalshi dropped fees as of 2026
NEAR_EXPIRY_HOURS   = 6      # trigger NearExpiry within this many hours of close
MIN_VOLUME          = 50     # skip markets with too little trading
CANDLE_INTERVAL_MIN = 60     # hourly candles (1, 60, or 1440)
LOOKBACK_BUFFER     = 2      # skip last N candles to avoid near-resolution data


# ═══════════════════════════════════════════════════════════════════════════
#  DATA CLASSES
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class CandleSignal:
    """A strategy signal fired at a specific candle."""
    ticker:      str
    title:       str
    strategy:    str
    candle_ts:   int      # Unix timestamp of the candle where signal fired
    entry_side:  str      # "yes" or "no"
    entry_price: float    # price at signal candle
    hours_left:  float    # hours until market close at signal time
    result:      str      # actual settlement result
    won:         bool
    net_pnl:     float
    volume_at_signal: float

@dataclass
class StrategyStats:
    name:   str
    trades: list = field(default_factory=list)

    @property
    def n(self): return len(self.trades)

    @property
    def wins(self): return sum(1 for t in self.trades if t.won)

    @property
    def win_rate(self): return self.wins / self.n if self.n else 0

    @property
    def total_pnl(self): return sum(t.net_pnl for t in self.trades)

    @property
    def avg_pnl(self): return self.total_pnl / self.n if self.n else 0

    @property
    def expected_value(self):
        total_risked = sum(t.entry_price for t in self.trades)
        return self.total_pnl / total_risked if total_risked else 0

    @property
    def profit_factor(self):
        w = sum(t.net_pnl for t in self.trades if t.net_pnl > 0)
        l = abs(sum(t.net_pnl for t in self.trades if t.net_pnl < 0))
        return w / l if l else float("inf")

    @property
    def avg_hours_left(self):
        return sum(t.hours_left for t in self.trades) / self.n if self.n else 0

    def summary_row(self):
        return [
            self.name,
            self.n,
            f"{self.win_rate:.1%}",
            f"${self.total_pnl:+.2f}",
            f"${self.avg_pnl:+.4f}",
            f"{self.expected_value:+.3f}",
            f"{self.profit_factor:.2f}x",
            f"{self.avg_hours_left:.1f}h",
        ]


# ═══════════════════════════════════════════════════════════════════════════
#  API CLIENT
# ═══════════════════════════════════════════════════════════════════════════

class KalshiClient:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers["Content-Type"] = "application/json"
        self.req_count = 0

    def _get(self, path, params=None, retries=3):
        url = f"{BASE_URL}{path}"
        for attempt in range(retries):
            try:
                r = self.session.get(url, params=params, timeout=15)
                self.req_count += 1
                if r.status_code == 429:
                    wait = int(r.headers.get("Retry-After", 10))
                    print(Fore.YELLOW + f"  Rate limited — waiting {wait}s...")
                    time.sleep(wait)
                    continue
                if r.status_code == 404:
                    return None
                r.raise_for_status()
                return r.json()
            except requests.HTTPError:
                return None
            except Exception as e:
                if attempt < retries - 1:
                    time.sleep(2)
                else:
                    print(Fore.RED + f"  Failed {path}: {e}")
                    return None

    def get_settled_markets(self, limit=200, debug=False):
        """
        Fetch settled markets with good candle history.
        Targets specific categories known to have long-duration markets:
        economics, politics, technology, climate — not short-lived sports.
        Falls back to general settled feed if categories yield too few.
        """
        # Categories that produce multi-hour/multi-day markets with good candle data
        GOOD_CATEGORIES = [
            "economics", "politics", "financials",
            "technology", "climate", "crypto",
            "pop culture", "geopolitics",
        ]
        markets = []
        seen    = set()

        def add_batch(raw, source=""):
            added = 0
            for m in raw:
                t = m.get("ticker","")
                if t in seen:
                    continue
                result = str(m.get("result","")).lower()
                if result not in ("yes","no"):
                    continue
                if m.get("mve_collection_ticker"):
                    continue
                if m.get("market_type") == "multivariate":
                    continue
                # Require at least 2 hours duration for meaningful candles
                try:
                    ot = datetime.fromisoformat(m["open_time"].replace("Z","+00:00")).timestamp()
                    ct = datetime.fromisoformat(m["close_time"].replace("Z","+00:00")).timestamp()
                    if (ct - ot) / 3600 < 2:
                        continue
                except:
                    pass
                seen.add(t)
                m["result"] = result
                markets.append(m)
                added += 1
            return added

        # ── Pass 1: fetch by category (best quality) ─────────────────────────
        for cat in GOOD_CATEGORIES:
            if len(markets) >= limit:
                break
            cursor = None
            for _ in range(5):
                params = {"limit": 200, "status": "settled", "category": cat}
                if cursor:
                    params["cursor"] = cursor
                data = self._get("/markets", params)
                if not data or not data.get("markets"):
                    break
                raw = data["markets"]
                if debug and not markets:
                    print(Fore.YELLOW + f"  [debug] category={cat} raw={len(raw)}")
                    if raw:
                        m0 = raw[0]
                        print(Fore.YELLOW + f"  [debug] fields: {sorted(m0.keys())}")
                        print(Fore.YELLOW + f"  [debug] result={m0.get('result')} vol={m0.get('volume_fp')} status={m0.get('status')}")
                added = add_batch(raw, cat)
                print(Fore.WHITE + f"  Category [{cat}]: +{added}  (total: {len(markets)})")
                cursor = data.get("cursor")
                if not cursor or len(markets) >= limit:
                    break
                time.sleep(0.2)

        # ── Pass 2: general settled + closed feed (fallback) ─────────────────
        if len(markets) < limit:
            for status in ["settled", "closed"]:
                if len(markets) >= limit:
                    break
                cursor = None
                pages  = 0
                while pages < max(5, limit // 40):
                    params = {"limit": 200, "status": status}
                    if cursor:
                        params["cursor"] = cursor
                    data = self._get("/markets", params)
                    if not data or not data.get("markets"):
                        break
                    raw = data["markets"]
                    added = add_batch(raw, status)
                    pages += 1
                    if added > 0:
                        print(Fore.WHITE + f"  General [{status}] page {pages}: +{added}  (total: {len(markets)})")
                    cursor = data.get("cursor")
                    if not cursor or len(markets) >= limit:
                        break
                    time.sleep(0.2)

        # ── Pass 3: historical endpoint ───────────────────────────────────────
        if len(markets) < limit:
            cursor = None
            for _ in range(10):
                params = {"limit": 200}
                if cursor:
                    params["cursor"] = cursor
                data = self._get("/historical/markets", params)
                if not data or not data.get("markets"):
                    break
                raw = data["markets"]
                added = add_batch(raw, "historical")
                if added > 0:
                    print(Fore.WHITE + f"  Historical: +{added}  (total: {len(markets)})")
                cursor = data.get("cursor")
                if not cursor or len(markets) >= limit:
                    break
                time.sleep(0.2)

        if not markets:
            print(Fore.RED + "  No suitable markets found. Try --debug for details.")

        print(Fore.CYAN + f"  Got {len(markets)} markets (2h+ duration, non-MVE).")
        return markets[:limit]

    @staticmethod
    def _derive_series_ticker(event_ticker):
        """
        Handles multiple Kalshi ticker formats:
          KXHIGHNY-23DEC31-T50       -> KXHIGHNY      (standard)
          KXMVECROSSCATEGORY-S2026XX -> KXMVECROSSCATEGORY (MVE)
          KXNFL-25JAN12              -> KXNFL          (sports)
        Strategy: take the FIRST hyphen-segment as the series.
        For standard markets that also matches the series.
        Falls back gracefully for any unknown format.
        """
        import re
        if not event_ticker:
            return ""
        parts = event_ticker.split("-")
        if not parts:
            return event_ticker
        # First part is always the series ticker base
        series = parts[0]
        # For some series with compound names like KX-SOMETHING, include more parts
        # but stop at any part that looks like a date (YYmmmDD) or hash/ID
        for part in parts[1:]:
            if re.match(r"^\d{2}[A-Z]{3}\d{2}", part):   # date segment
                break
            if re.match(r"^[A-Z]\d{4}", part):              # ID segment like S2026X
                break
            if len(part) > 8 and re.search(r"[0-9]{4}", part):  # long hash-like
                break
            series += "-" + part
        return series

    def get_candlesticks(self, ticker, event_ticker, series_ticker,
                         start_ts, end_ts, period_interval=60, debug=False):
        """
        Fetch candlesticks — tries 5 endpoint patterns with fallback.
        """
        params_base = {
            "start_ts":        int(start_ts),
            "end_ts":          int(end_ts),
            "period_interval": period_interval,
        }
        derived = self._derive_series_ticker(event_ticker) or series_ticker

        def extract(data):
            if not data:
                return None
            if data.get("market_candlesticks"):
                arrays = data["market_candlesticks"]
                tlist  = data.get("market_tickers", [])
                # Try to match our specific ticker
                for i, t in enumerate(tlist):
                    if t == ticker and i < len(arrays):
                        arr = [c for c in arrays[i] if c]  # filter nulls
                        if arr:
                            return arr
                # Fall back to first non-empty array
                for arr in arrays:
                    clean = [c for c in arr if c]
                    if clean:
                        return clean
            if data.get("candlesticks"):
                return [c for c in data["candlesticks"] if c]
            return None

        attempts = []
        if derived and event_ticker:
            attempts.append((f"/series/{derived}/events/{event_ticker}/candlesticks", {}))
        attempts.append(("/markets/candlesticks", {"tickers": ticker}))
        attempts.append((f"/markets/{ticker}/candlesticks", {}))
        if derived and event_ticker:
            attempts.append((f"/historical/series/{derived}/events/{event_ticker}/candlesticks", {}))
        attempts.append((f"/historical/markets/{ticker}/candlesticks", {}))

        for path, extra in attempts:
            params = {**params_base, **extra}
            if debug:
                print(f"    trying: {path}")
            data = self._get(path, params)
            candles = extract(data)
            if candles:
                if debug:
                    print(f"    got {len(candles)} candles")
                return candles
            elif debug and data is not None:
                print(f"    response keys: {list(data.keys())}")
        return None


# ═══════════════════════════════════════════════════════════════════════════
#  SIGNAL FUNCTIONS  (operate on a single candle + context)
# ═══════════════════════════════════════════════════════════════════════════

def _p(candle, path):
    """Safely extract a nested dollar price from candle."""
    try:
        parts = path.split(".")
        obj = candle
        for p in parts:
            obj = obj[p]
        return float(obj)
    except:
        return None

def signal_spread_gap(candle, hours_left, cumulative_volume):
    """
    YES bid + NO bid < 0.97 mid-market.
    In a binary market: NO bid ≈ 1 - YES ask
    So gap = YES ask - YES bid  (the spread).
    We flag when: yes_bid + (1 - yes_ask) < 0.97
    i.e. yes_ask - yes_bid > 0.03  (wide spread)
    """
    yes_bid = _p(candle, "yes_bid.close_dollars")
    yes_ask = _p(candle, "yes_ask.close_dollars")
    if yes_bid is None or yes_ask is None:
        return False, None, None
    if cumulative_volume < MIN_VOLUME:
        return False, None, None

    no_bid = 1.0 - yes_ask   # in binary market
    total  = yes_bid + no_bid
    if total < 0.97:
        side  = "yes" if yes_bid <= no_bid else "no"
        price = yes_bid if side == "yes" else no_bid
        return True, side, price
    return False, None, None

def signal_near_expiry(candle, hours_left, cumulative_volume):
    """
    Market is within NEAR_EXPIRY_HOURS of close AND price is extreme.
    Extreme = YES bid < 0.08 (bet NO) or YES ask > 0.92 (bet YES).
    Exclude the very last candle window to avoid look-ahead.
    """
    if hours_left > NEAR_EXPIRY_HOURS or hours_left < 0.5:
        return False, None, None
    if cumulative_volume < MIN_VOLUME:
        return False, None, None

    yes_bid = _p(candle, "yes_bid.close_dollars")
    yes_ask = _p(candle, "yes_ask.close_dollars")
    if yes_bid is None:
        return False, None, None

    no_bid = 1.0 - (yes_ask or 1.0 - yes_bid)

    if yes_bid < 0.08:   # market strongly pricing NO
        return True, "no", no_bid
    if yes_ask is not None and yes_ask > 0.92:  # market strongly pricing YES
        return True, "yes", yes_ask
    return False, None, None

def signal_momentum(candle, hours_left, cumulative_volume):
    """
    High-volume market trending hard toward YES or NO resolution.
    More stringent than NearExpiry — requires high volume AND extreme price
    AND not too close to expiry (want time to confirm).
    """
    if cumulative_volume < 500:
        return False, None, None
    if hours_left < 1.0:   # too close — likely look-ahead territory
        return False, None, None

    yes_bid = _p(candle, "yes_bid.close_dollars")
    yes_ask = _p(candle, "yes_ask.close_dollars")
    if yes_bid is None:
        return False, None, None

    no_bid = 1.0 - (yes_ask or 1.0 - yes_bid)

    if yes_bid > 0.85:
        return True, "yes", yes_bid
    if yes_bid < 0.15:
        return True, "no", no_bid
    return False, None, None


STRATEGIES = {
    "SpreadGap":  signal_spread_gap,
    "NearExpiry": signal_near_expiry,
    "Momentum":   signal_momentum,
}


# ═══════════════════════════════════════════════════════════════════════════
#  P&L
# ═══════════════════════════════════════════════════════════════════════════

def compute_pnl(side, entry_price, result):
    gross = (1.0 - entry_price) if side == result else (-entry_price)
    return gross - FEE_PER_CONTRACT


# ═══════════════════════════════════════════════════════════════════════════
#  BACKTEST ENGINE
# ═══════════════════════════════════════════════════════════════════════════

def backtest_market(client, market, strategies, verbose=False):
    """
    Fetch candles for one market, walk them in time order,
    fire signals without looking ahead, record first trigger per strategy.
    Returns list of CandleSignal.
    """
    ticker        = market["ticker"]
    result        = market.get("result")
    event_ticker  = market.get("event_ticker", "")
    series_ticker = market.get("series_ticker", "")
    title         = market.get("title", "")[:60]

    # Parse market open/close times
    try:
        open_ts  = int(datetime.fromisoformat(
            market["open_time"].replace("Z", "+00:00")).timestamp())
        close_ts = int(datetime.fromisoformat(
            market["close_time"].replace("Z", "+00:00")).timestamp())
    except:
        return []

    # Skip markets shorter than 30 minutes
    duration_hours = (close_ts - open_ts) / 3600
    if verbose:
        print(Fore.WHITE + f"    duration={duration_hours:.1f}h  event={event_ticker}")
    if duration_hours < 0.5:
        if verbose:
            print(Fore.YELLOW + f"    skipped: too short ({duration_hours:.1f}h)")
        return []

    candles = client.get_candlesticks(
        ticker=ticker,
        event_ticker=event_ticker,
        series_ticker=series_ticker,
        start_ts=open_ts,
        end_ts=close_ts,
        period_interval=CANDLE_INTERVAL_MIN,
        debug=verbose,
    )

    if not candles or len(candles) < 2:
        return []

    # Sort candles by timestamp
    candles = sorted(candles, key=lambda c: c.get("end_period_ts", 0))

    # Drop last LOOKBACK_BUFFER candles to avoid near-resolution data
    safe_candles = candles[:-LOOKBACK_BUFFER] if len(candles) > LOOKBACK_BUFFER else candles[:-1]
    if not safe_candles:
        return []

    signals = []
    triggered = {name: False for name in strategies}
    cumulative_volume = 0.0

    for candle in safe_candles:
        candle_ts = candle.get("end_period_ts", 0)
        hours_left = (close_ts - candle_ts) / 3600
        vol = float(candle.get("volume_fp", 0) or 0)
        cumulative_volume += vol

        for strat_name, signal_fn in strategies.items():
            if triggered[strat_name]:
                continue  # only take first signal per market per strategy

            fired, side, entry_price = signal_fn(candle, hours_left, cumulative_volume)
            if not fired or side is None or entry_price is None:
                continue
            if entry_price <= 0 or entry_price >= 1.0:
                continue

            net_pnl = compute_pnl(side, entry_price, result)
            won     = (side == result)

            sig = CandleSignal(
                ticker=ticker,
                title=title,
                strategy=strat_name,
                candle_ts=candle_ts,
                entry_side=side,
                entry_price=entry_price,
                hours_left=hours_left,
                result=result,
                won=won,
                net_pnl=net_pnl,
                volume_at_signal=cumulative_volume,
            )
            signals.append(sig)
            triggered[strat_name] = True

            if verbose:
                ts_str = datetime.fromtimestamp(candle_ts).strftime("%m/%d %H:%M")
                icon = "✓" if won else "✗"
                print(f"    {icon} {strat_name:12} {ticker[:28]:28} "
                      f"{side.upper():3} @ ${entry_price:.2f}  "
                      f"{hours_left:.1f}h left  "
                      f"result={result.upper()}  P&L=${net_pnl:+.4f}  [{ts_str}]")

    return signals


def run_backtest(client, markets, strategy_names=None, verbose=False):
    if strategy_names is None:
        strategy_names = list(STRATEGIES.keys())

    selected = {k: STRATEGIES[k] for k in strategy_names if k in STRATEGIES}
    stats    = {name: StrategyStats(name=name) for name in strategy_names}
    all_signals = []
    skipped = 0

    for i, market in enumerate(markets):
        ticker = market.get("ticker", "?")
        if verbose:
            print(Fore.WHITE + f"\n  [{i+1}/{len(markets)}] {ticker}")

        sigs = backtest_market(client, market, selected, verbose=verbose)
        if not sigs:
            skipped += 1
        for sig in sigs:
            stats[sig.strategy].trades.append(sig)
            all_signals.append(sig)

        # Progress indicator (non-verbose)
        if not verbose and (i + 1) % 10 == 0:
            wins  = sum(1 for s in all_signals if s.won)
            total = len(all_signals)
            wr    = f"{wins/total:.0%}" if total else "—"
            print(Fore.WHITE + f"  [{i+1}/{len(markets)}] signals={total}  "
                  f"overall_win_rate={wr}  api_calls={client.req_count}")

        time.sleep(0.15)  # polite rate limiting

    print(Fore.WHITE + f"\n  Markets with no candle data: {skipped}")
    return stats, all_signals


# ═══════════════════════════════════════════════════════════════════════════
#  DISPLAY
# ═══════════════════════════════════════════════════════════════════════════

def print_summary(stats, total_markets):
    print(Fore.CYAN + f"\n{'═'*105}")
    print(Fore.CYAN + f"  KALSHI CANDLESTICK BACKTEST  —  {total_markets} markets  "
          f"(look-ahead bias eliminated)")
    print(Fore.CYAN + f"{'═'*105}\n")

    headers = ["Strategy", "Trades", "Win Rate", "Total P&L", "Avg P&L",
               "EV/$1 Risked", "Profit Factor", "Avg Entry Hours Left"]
    rows = [s.summary_row() for s in stats.values()]
    print(tabulate(rows, headers=headers, tablefmt="rounded_outline"))
    print()

    for name, s in stats.items():
        if s.n == 0:
            print(Fore.YELLOW + f"  {name}: No signals triggered. "
                  f"Try --markets with a larger number.")
            continue
        color = Fore.GREEN if s.total_pnl > 0 else Fore.RED
        verdict = "PROFITABLE ✓" if s.expected_value > 0.02 else \
                  "MARGINAL"   if s.expected_value > 0 else "UNPROFITABLE ✗"
        print(color + f"  {name}: {s.wins}/{s.n} wins | "
              f"${s.total_pnl:+.2f} P&L | EV={s.expected_value:+.3f} | {verdict}")

        # Category breakdown
        by_cat = defaultdict(lambda: [0, 0, 0.0])
        for t in s.trades:
            cat = "unknown"
            by_cat[cat][0] += 1
            by_cat[cat][1] += int(t.won)
            by_cat[cat][2] += t.net_pnl
        cat_rows = sorted(
            [[c, v[0], f"{v[1]/v[0]:.1%}", f"${v[2]:+.2f}"]
             for c, v in by_cat.items()],
            key=lambda r: r[0]
        )
        print(tabulate(cat_rows,
                       headers=["  Category", "Trades", "Win%", "P&L"],
                       tablefmt="plain",
                       colalign=("left", "right", "right", "right")))
        print()


def print_sample_trades(signals, n=15):
    if not signals:
        return
    print(Fore.CYAN + f"  Sample trades (showing {min(n, len(signals))} of {len(signals)}):\n")
    rows = []
    for s in signals[:n]:
        ts = datetime.fromtimestamp(s.candle_ts).strftime("%m/%d %H:%M")
        rows.append([
            s.strategy,
            s.ticker[:26],
            s.entry_side.upper(),
            f"${s.entry_price:.2f}",
            f"{s.hours_left:.1f}h",
            s.result.upper(),
            "✓ WIN" if s.won else "✗ LOSS",
            f"${s.net_pnl:+.4f}",
            ts,
        ])
    print(tabulate(rows,
                   headers=["Strategy", "Ticker", "Side", "Entry",
                             "Hrs Left", "Result", "Outcome", "P&L", "Signal Time"],
                   tablefmt="rounded_outline"))
    print()


def print_hours_left_breakdown(stats):
    """Show win rate by how many hours were left when signal fired."""
    print(Fore.CYAN + "  WIN RATE BY HOURS-TO-EXPIRY AT SIGNAL TIME:\n")
    buckets = [(0, 1), (1, 3), (3, 6), (6, 12), (12, 48), (48, 9999)]
    for name, s in stats.items():
        if s.n < 5:
            continue
        rows = []
        for lo, hi in buckets:
            trades = [t for t in s.trades if lo <= t.hours_left < hi]
            if not trades:
                continue
            wr  = sum(1 for t in trades if t.won) / len(trades)
            pnl = sum(t.net_pnl for t in trades)
            label = f"{lo}–{hi}h" if hi < 9999 else f"{lo}h+"
            rows.append([label, len(trades), f"{wr:.1%}", f"${pnl:+.2f}"])
        if rows:
            print(Fore.WHITE + f"  {name}:")
            print(tabulate(rows,
                           headers=["  Hours Left", "N", "Win%", "P&L"],
                           tablefmt="plain",
                           colalign=("left", "right", "right", "right")))
            print()


# ═══════════════════════════════════════════════════════════════════════════
#  SAVE TO EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def save_excel(stats, all_signals, filename="kalshi_backtest_v2_results.xlsx"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb  = openpyxl.Workbook()
        hfill = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True)

        def style_header(ws, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col)].width = 18

        # Summary
        ws = wb.active
        ws.title = "Summary"
        style_header(ws, ["Strategy","Trades","Win Rate","Total P&L",
                           "Avg P&L","EV per $1","Profit Factor","Avg Hrs Left"])
        for r, s in enumerate(stats.values(), 2):
            for c, v in enumerate(s.summary_row(), 1):
                ws.cell(row=r, column=c, value=v)

        # All trades
        wt = wb.create_sheet("All Trades")
        style_header(wt, ["Strategy","Ticker","Title","Side","Entry Price",
                           "Hours Left","Result","Won","Net P&L",
                           "Volume at Signal","Signal Time"])
        for r, s in enumerate(all_signals, 2):
            ts = datetime.fromtimestamp(s.candle_ts).strftime("%Y-%m-%d %H:%M")
            vals = [s.strategy, s.ticker, s.title, s.entry_side.upper(),
                    round(s.entry_price, 4), round(s.hours_left, 1),
                    s.result.upper(), "WIN" if s.won else "LOSS",
                    round(s.net_pnl, 4), round(s.volume_at_signal, 1), ts]
            for c, v in enumerate(vals, 1):
                wt.cell(row=r, column=c, value=v)

        wb.save(filename)
        print(Fore.GREEN + f"\n  Results saved to {filename}")
    except ImportError:
        print(Fore.YELLOW + "  pip install openpyxl to enable Excel export")
    except Exception as e:
        print(Fore.RED + f"  Excel save failed: {e}")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Kalshi Candlestick Backtester v2")
    parser.add_argument("--markets",  type=int, default=100,
                        help="Number of settled markets to test (default: 100)")
    parser.add_argument("--strategy", type=str, default=None,
                        help="Test one strategy: SpreadGap | NearExpiry | Momentum")
    parser.add_argument("--save",     action="store_true",
                        help="Save results to Excel")
    parser.add_argument("--verbose",  action="store_true",
                        help="Show every signal as it fires")
    parser.add_argument("--debug",    action="store_true",
                        help="Show candlestick API endpoint attempts (diagnose issues)")
    args = parser.parse_args()

    strats = ([args.strategy] if args.strategy and args.strategy in STRATEGIES
              else list(STRATEGIES.keys()))

    print(Fore.CYAN + f"\n{'═'*65}")
    print(Fore.CYAN + f"  KALSHI CANDLESTICK BACKTESTER v2")
    print(Fore.CYAN + f"  Strategies : {', '.join(strats)}")
    print(Fore.CYAN + f"  Markets    : {args.markets}")
    print(Fore.CYAN + f"  Candle res : {CANDLE_INTERVAL_MIN}m  "
          f"(buffer: skip last {LOOKBACK_BUFFER} candles)")
    print(Fore.CYAN + f"  Fees       : ${FEE_PER_CONTRACT:.2f}/contract")
    print(Fore.CYAN + f"{'═'*65}\n")

    client = KalshiClient()

    print(Fore.WHITE + "  Step 1: Fetching settled markets...")
    markets = client.get_settled_markets(limit=args.markets, debug=args.debug)
    if not markets:
        print(Fore.RED + "  No markets found — check your connection.")
        return
    print(Fore.CYAN + f"  Got {len(markets)} markets with known outcomes.\n")

    print(Fore.WHITE + "  Step 2: Fetching candlestick history & running backtest...")
    print(Fore.WHITE + f"  (This makes ~{len(markets)} API calls — may take a minute)\n")
    stats, all_signals = run_backtest(client, markets,
                                      strategy_names=strats,
                                      verbose=args.verbose or args.debug)

    print(Fore.WHITE + "\n  Step 3: Results\n")
    print_summary(stats, len(markets))
    print_sample_trades(all_signals)
    print_hours_left_breakdown(stats)

    if args.save:
        save_excel(stats, all_signals)

    print(Fore.WHITE + f"  Total API calls: {client.req_count}")
    print(Fore.YELLOW + "\n  NOTE: Each signal is the FIRST trigger per market "
          "per strategy, using only data available at that candle.")
    print(Fore.YELLOW + "  Win rates above 55% with EV > 0.02 are worth "
          "live-testing on demo with small size.")


if __name__ == "__main__":
    main()
